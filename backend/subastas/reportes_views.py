"""
Views para Reportes de Subastas.

Genera reportes profesionales en formato Excel con información completa
de las subastas: empresa, tipo de fruta, ganador, ofertas, etc.
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Count, Max, Min
from django.utils import timezone
from usuarios.permissions import RBACPermission
from datetime import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Subasta, Oferta
from clientes.models import Cliente
from modulo_packing.models import PackingSemanal, PackingDetalle


class ReporteSubastasViewSet(viewsets.ViewSet):
    """
    ViewSet para generar reportes de subastas.
    
    Endpoints:
    - GET /api/admin/reportes/subastas/excel/ - Generar reporte Excel
    """
    
    permission_classes = [IsAuthenticated, RBACPermission]
    modulo_permiso = 'reportes'
    
    permisos_mapping = {
        'excel': 'generate_auctions',
        'clientes_excel': 'generate_clients',
        'packing_excel': 'generate_packings',
    }
    
    def _formato_semana(self, fecha_inicio, fecha_fin):
        """
        Formatea el rango de fechas al formato:
        'Semana {num_semana} | {dd}/{mm}/{yyyy} - {dd}/{mm}/{yyyy}'
        Ejemplo: Semana 8 | 16/02/2026 - 21/02/2026
        """
        semana = fecha_inicio.isocalendar()[1]
        
        return f"Semana {semana} | {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    
    @action(detail=False, methods=['get'])
    def excel(self, request):
        """
        Genera un reporte Excel profesional de subastas.
        
        Parámetros query:
        - fecha_inicio: Fecha de inicio del rango (formato: YYYY-MM-DD)
        - fecha_fin: Fecha de fin del rango (formato: YYYY-MM-DD)
        
        Retorna: Archivo Excel descargable
        """
        # Obtener parámetros de filtro
        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)
        
        # Validar que fecha_inicio sea obligatoria
        if not fecha_inicio_str:
            return Response(
                {'error': 'La fecha de inicio es obligatoria para generar el reporte.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Construir queryset base con todas las relaciones necesarias
        queryset = Subasta.objects.select_related(
            'packing_detalle',
            'packing_detalle__packing_tipo',
            'packing_detalle__packing_tipo__tipo_fruta',
            'packing_detalle__packing_tipo__packing_semanal',
            'packing_detalle__packing_tipo__packing_semanal__empresa',
        ).prefetch_related('ofertas', 'ofertas__cliente').order_by(
            '-packing_detalle__packing_tipo__packing_semanal__fecha_inicio_semana',  # Semana descendente (más recientes primero)
            'packing_detalle__fecha',  # Día ascendente (Lunes → Sábado)
            'packing_detalle__packing_tipo__tipo_fruta__nombre'  # Tipo de fruta alfabéticamente
        )
        
        # Aplicar filtros de fecha si se proporcionan
        if fecha_inicio_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_hora_inicio__date__gte=fecha_inicio)
            except ValueError:
                return Response(
                    {'error': 'Formato de fecha_inicio inválido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        if fecha_fin_str:
            try:
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_hora_inicio__date__lte=fecha_fin)
            except ValueError:
                return Response(
                    {'error': 'Formato de fecha_fin inválido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Generar el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Subastas"
        
        # =====================================================================
        # ESTILOS
        # =====================================================================
        
        # Estilo del título
        titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        titulo_alignment = Alignment(horizontal='center', vertical='center')
        
        # Estilo del encabezado de columnas
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Estilo de las celdas de datos
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        cell_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # =====================================================================
        # TÍTULO DEL REPORTE
        # =====================================================================
        
        ws.merge_cells('A1:O1')
        cell_titulo = ws['A1']
        
        # Crear texto del título con rango de fechas
        titulo_texto = "REPORTE DE SUBASTAS"
        if fecha_inicio_str or fecha_fin_str:
            if fecha_inicio_str and fecha_fin_str:
                titulo_texto += f" - Del {fecha_inicio_str} al {fecha_fin_str}"
            elif fecha_inicio_str:
                titulo_texto += f" - Desde {fecha_inicio_str}"
            else:
                titulo_texto += f" - Hasta {fecha_fin_str}"
        
        cell_titulo.value = titulo_texto
        cell_titulo.font = titulo_font
        cell_titulo.fill = titulo_fill
        cell_titulo.alignment = titulo_alignment
        ws.row_dimensions[1].height = 30
        
        # =====================================================================
        # ENCABEZADOS DE COLUMNAS
        # =====================================================================
        
        headers = [
            'ID Subasta',
            'Empresa',
            'Semana de Proyectado',
            'Día',
            'Fecha Producción',
            'Tipo Fruta',
            'Kilos',
            'Estado',
            'Fecha Inicio',
            'Fecha Fin',
            'Duración',
            'Precio Base',
            'Ganador',
            'Monto Ganador',
            'Total Ofertas'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            
            # COLORES DIFERENCIADOS PARA CABECERAS
            # Datos de Packing (Cols 2-7): Naranja
            if 2 <= col_num <= 7:
                cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
            # Datos de Subasta (Cols 8-14): Azul
            elif 8 <= col_num <= 14:
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            else:
                cell.fill = header_fill # Default (ID Subasta)
            cell.alignment = header_alignment
            cell.border = cell_border
        
        ws.row_dimensions[2].height = 35
        
        # =====================================================================
        # DATOS DE SUBASTAS
        # =====================================================================
        
        # Obtener tiempo actual
        ahora = timezone.now()
        
        row_num = 3
        for subasta in queryset:
            # 1. Calcular estado en tiempo real (Lógica dinámica)
            estado_real = subasta.estado
            if subasta.estado != 'CANCELADA':
                if ahora < subasta.fecha_hora_inicio:
                    estado_real = 'PROGRAMADA'
                elif subasta.fecha_hora_inicio <= ahora <= subasta.fecha_hora_fin:
                    estado_real = 'ACTIVA'
                else:
                    estado_real = 'FINALIZADA'

            # 2. Obtener información de la oferta ganadora (Optimizado en memoria)
            # Usamos las ofertas ya pre-cargadas en subasta.ofertas.all()
            # en lugar de subasta.oferta_ganadora que hace una nueva consulta.
            ofertas = list(subasta.ofertas.all())
            oferta_ganadora = None
            if ofertas:
                # Encontrar la de mayor monto (y luego fecha más reciente si hay empate, 
                # aunque lógica de negocio dice que monto debe ser mayor)
                oferta_ganadora = max(ofertas, key=lambda o: o.monto)
                # Validar si realmente es ganadora (flag es_ganadora)
                if not oferta_ganadora.es_ganadora:
                     # Si por alguna razón la de mayor monto no está marcada (error de consistencia?)
                     # buscamos la que tenga es_ganadora=True
                     ganadoras = [o for o in ofertas if o.es_ganadora]
                     if ganadoras:
                         oferta_ganadora = ganadoras[0]
                     else:
                         oferta_ganadora = None

            # 3. Determinar qué mostrar como ganador según el estado dinámico
            if estado_real == 'FINALIZADA':
                ganador_nombre = oferta_ganadora.cliente.nombre_razon_social if oferta_ganadora else 'Sin ofertas'
                monto_ganador = float(oferta_ganadora.monto) if oferta_ganadora else 0.0
            elif estado_real == 'ACTIVA':
                ganador_nombre = f"LÍDER: {oferta_ganadora.cliente.nombre_razon_social}" if oferta_ganadora else 'Sin ofertas'
                monto_ganador = float(oferta_ganadora.monto) if oferta_ganadora else 0.0
            elif estado_real == 'CANCELADA':
                ganador_nombre = 'CANCELADA'
                monto_ganador = 0.0
            else: # PROGRAMADA
                ganador_nombre = 'Pendiente'
                monto_ganador = 0.0
            
            # Contar ofertas totales
            total_ofertas = subasta.ofertas.count()
            
            # Obtener datos del packing
            detalle = subasta.packing_detalle
            packing_tipo = detalle.packing_tipo
            packing_semanal = packing_tipo.packing_semanal
            
            # Convertir horarios UTC a zona horaria de Perú (America/Lima, UTC-5)
            lima_tz = pytz.timezone('America/Lima')
            fecha_inicio_peru = subasta.fecha_hora_inicio.astimezone(lima_tz)
            fecha_fin_peru = subasta.fecha_hora_fin.astimezone(lima_tz)
            
            # Calcular duración de la subasta
            duracion_delta = subasta.fecha_hora_fin - subasta.fecha_hora_inicio
            duracion_segundos = int(duracion_delta.total_seconds())
            duracion_horas = duracion_segundos // 3600
            duracion_minutos = (duracion_segundos % 3600) // 60
            duracion_str = f"{duracion_horas}h {duracion_minutos}m"
            
            # Datos de la fila
            row_data = [
                subasta.id,
                packing_semanal.empresa.nombre,
                self._formato_semana(packing_semanal.fecha_inicio_semana, packing_semanal.fecha_fin_semana),
                detalle.get_dia_display(),
                detalle.fecha.strftime('%d/%m/%Y'),
                packing_tipo.tipo_fruta.nombre,
                float(detalle.py),
                estado_real, # Usamos el estado calculado en tiempo real
                fecha_inicio_peru.strftime('%d/%m/%Y %H:%M'),  # Horario de Perú
                fecha_fin_peru.strftime('%d/%m/%Y %H:%M'),      # Horario de Perú
                duracion_str,  # Nueva columna de duración
                float(subasta.precio_base),
                ganador_nombre,
                monto_ganador,
                total_ofertas
            ]
            
            # Escribir la fila
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = cell_border
                
                # Formato especial para números
                if col_num in [7, 12, 14]:  # Kilos, Precio Base, Monto Ganador
                    cell.number_format = '#,##0.00'
            
            row_num += 1
        
        # =====================================================================
        # AJUSTAR ANCHOS DE COLUMNAS
        # =====================================================================
        
        column_widths = {
            'A': 12,  # ID Subasta
            'B': 25,  # Empresa
            'C': 35,  # Semana (Aumentado para mostrar "Semana X | dd/mm/yyyy...")
            'D': 12,  # Día
            'E': 18,  # Fecha Producción
            'F': 20,  # Tipo Fruta
            'G': 12,  # Kilos
            'H': 15,  # Estado
            'I': 18,  # Fecha Inicio
            'J': 18,  # Fecha Fin
            'K': 12,  # Duración
            'L': 14,  # Precio Base
            'M': 30,  # Ganador
            'N': 15,  # Monto Ganador
            'O': 14,  # Total Ofertas
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # =====================================================================
        # AGREGAR FILTROS AUTOMÁTICOS
        # =====================================================================
        
        if row_num > 3:  # Solo si hay datos
            ws.auto_filter.ref = f"A2:O{row_num-1}"
        
        # =====================================================================
        # GENERAR RESPUESTA HTTP CON EL ARCHIVO
        # =====================================================================
        
        # Crear nombre de archivo con timestamp
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"reporte_subastas_{timestamp}.xlsx"
        
        # Preparar la respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar el workbook en la respuesta
        wb.save(response)
        
        return response
    
    @action(detail=False, methods=['get'])
    def clientes_excel(self, request):
        """
        Genera un reporte Excel profesional de clientes.
        
        Incluye información completa del cliente y estadísticas de participación.
        
        Retorna: Archivo Excel descargable
        """
        from django.db.models import Count, Sum, Q, Avg, Max
        
        # Obtener todos los clientes con sus estadísticas
        clientes = Cliente.objects.all().order_by('nombre_razon_social')
        
        # Generar el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Clientes"
        
        # =====================================================================
        # ESTILOS
        # =====================================================================
        
        # Estilo del título
        titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        titulo_alignment = Alignment(horizontal='center', vertical='center')
        
        # Estilo del encabezado de columnas
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Estilo de las celdas de datos
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        cell_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # =====================================================================
        # TÍTULO DEL REPORTE
        # =====================================================================
        
        ws.merge_cells('A1:V1')
        cell_titulo = ws['A1']
        cell_titulo.value = "REPORTE DE CLIENTES"
        cell_titulo.font = titulo_font
        cell_titulo.fill = titulo_fill
        cell_titulo.alignment = titulo_alignment
        ws.row_dimensions[1].height = 30
        
        # =====================================================================
        # ENCABEZADOS DE COLUMNAS
        # =====================================================================
        
        headers = [
            'RUC/DNI',
            'Fecha Registro',
            'Nombre / Razón Social',
            'Tipo',
            'Sede',
            'Estado',
            'Contacto 1',
            'Cargo 1',
            'Teléfono 1',
            'Email 1',
            'Contacto 2',
            'Cargo 2',
            'Teléfono 2',
            'Email 2',
            'Estatus Ficha',
            'Correo Conf.',
            'Participación (Subastas)',
            'Subastas Ganadas',
            'Subastas Perdidas',
            'Subastas en Curso',
            'Total Ofertas',
            'Creado Por'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
        
        ws.row_dimensions[2].height = 35
        
        # =====================================================================
        # DATOS DE CLIENTES
        # =====================================================================
        
        # Obtener tiempo actual para cálculos real-time
        ahora = timezone.now()
        
        row_num = 3
        for cliente in clientes:
            # Base queryset para ofertas del cliente en subastas NO canceladas
            ofertas_validas = Oferta.objects.filter(cliente=cliente).exclude(subasta__estado='CANCELADA')
            
            # IDs de subastas donde ha participado (activas o finalizadas, NO canceladas)
            ids_subastas_participadas = ofertas_validas.values_list('subasta_id', flat=True).distinct()
            subastas_validadas = Subasta.objects.filter(id__in=ids_subastas_participadas).exclude(estado='CANCELADA')
            
            # Desglose por tiempo
            finalizadas_ids = subastas_validadas.filter(fecha_hora_fin__lt=ahora).values_list('id', flat=True)
            activas_ids = subastas_validadas.filter(fecha_hora_inicio__lte=ahora, fecha_hora_fin__gte=ahora).values_list('id', flat=True)
            
            # Subastas ganadas: Oferta ganadora en subasta que ya terminó
            subastas_ganadas = ofertas_validas.filter(
                subasta_id__in=finalizadas_ids,
                es_ganadora=True
            ).count()
            
            # Subastas perdidas: Finalizadas donde no ganó
            subastas_perdidas = len(finalizadas_ids) - subastas_ganadas
            
            # Subastas en curso
            subastas_en_curso = len(activas_ids)
            
            # Métricas monetarias
            # Métricas monetarias REMOVIDAS a petición
            # agg = ofertas_validas.aggregate(promedio=Avg('monto'), maximo=Max('monto'))
            
            # Obtener creador (compatibilidad con `PerfilUsuario` que envuelve a `User`)
            creador = "Sistema/Desconocido"
            perfil = getattr(cliente, 'creado_por', None)
            if perfil:
                # PerfilUsuario tiene un OneToOne `user` con los campos reales
                user = getattr(perfil, 'user', None)
                if user:
                    nombre_completo = " ".join(filter(None, [getattr(user, 'first_name', ''), getattr(user, 'last_name', '')])).strip()
                    creador = nombre_completo or getattr(user, 'username', None) or str(perfil)
                else:
                    creador = str(perfil)

            # Datos de la fila
            row_data = [
                cliente.ruc_dni,
                cliente.fecha_creacion.strftime('%d/%m/%Y'),
                cliente.nombre_razon_social,
                cliente.get_tipo_display(),
                cliente.sede,
                cliente.get_estado_display(),
                cliente.contacto_1,
                cliente.cargo_1,
                cliente.numero_1,
                cliente.correo_electronico_1,
                cliente.contacto_2 or '',
                cliente.cargo_2 or '',
                cliente.numero_2 or '',
                cliente.correo_electronico_2 or '',
                cliente.get_estatus_ficha_display(),
                'Sí' if cliente.confirmacion_correo else 'No',
                subastas_validadas.count(),  # Participación
                subastas_ganadas,
                subastas_perdidas,
                subastas_en_curso,
                ofertas_validas.count(),     # Total Ofertas
                creador
            ]
            
            # Escribir la fila
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = cell_border
            
            row_num += 1
        
        # =====================================================================
        # AJUSTAR ANCHOS DE COLUMNAS
        # =====================================================================
        
        column_widths = {
            'A': 15,  # RUC/DNI
            'B': 18,  # Fecha Registro (MOVIDO)
            'C': 35,  # Nombre / Razón Social
            'D': 18,  # Tipo
            'E': 20,  # Sede
            'F': 15,  # Estado
            'G': 25,  # Contacto 1
            'H': 20,  # Cargo 1
            'I': 15,  # Teléfono 1
            'J': 30,  # Email 1
            'K': 25,  # Contacto 2
            'L': 20,  # Cargo 2
            'M': 15,  # Teléfono 2
            'N': 30,  # Email 2
            'O': 18,  # Estatus Ficha
            'P': 14,  # Correo Conf.
            'Q': 16,  # Participación
            'R': 16,  # Subastas Ganadas
            'S': 16,  # Subastas Perdidas
            'T': 16,  # Subastas en Curso
            'U': 15,  # Total Ofertas
            'V': 25,  # Creado Por (NUEVO)
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # =====================================================================
        # AGREGAR FILTROS AUTOMÁTICOS
        # =====================================================================
        
        if row_num > 3:  # Solo si hay datos
            ws.auto_filter.ref = f"A2:V{row_num-1}"
            
            # Formato moneda eliminado
        
        
        # =====================================================================
        # GENERAR RESPUESTA HTTP CON EL ARCHIVO
        # =====================================================================
        
        # Crear nombre de archivo con timestamp
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"reporte_clientes_{timestamp}.xlsx"
        
        # Preparar la respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar el workbook en la respuesta
        wb.save(response)
        
        return response

    @action(detail=False, methods=['get'])
    def packing_excel(self, request):
        """
        Genera un reporte Excel profesional de packing (producción).
        
        Incluye desglose diario, totales por tipo y estado.
        """
        # Parámetros de filtro
        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)
        
        # Validar que fecha_inicio sea obligatoria
        if not fecha_inicio_str:
            return Response(
                {'error': 'La fecha de inicio es obligatoria para generar el reporte.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Queryset base
        queryset = PackingSemanal.objects.select_related('empresa').prefetch_related(
            'tipos', 
            'tipos__tipo_fruta', 
            'tipos__detalles',
            'tipos__detalles__subastas'  # PREFETCH CRÍTICO: Evita N+1 en el bucle (plural por ForeignKey)
        ).order_by('-fecha_inicio_semana', 'empresa__nombre')
        
        # Filtros de fecha
        if fecha_inicio_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_inicio_semana__gte=fecha_inicio)
            except ValueError:
                return Response({'error': 'Formato fecha_inicio inválido'}, status=400)
                
        if fecha_fin_str:
            try:
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_inicio_semana__lte=fecha_fin)
            except ValueError:
                return Response({'error': 'Formato fecha_fin inválido'}, status=400)

        # Generar Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Packing"
        
        # Estilos (Estética Naranja Profesional)
        titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='E46C0A', end_color='E46C0A', fill_type='solid') # Naranja para Packing
        titulo_alignment = Alignment(horizontal='center', vertical='center')
        
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        cell_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # --- DEFINIR COLORES DE ESTADO PARA CELDAS DIARIAS ---
        fill_finalizada = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid') # Azul claro
        fill_activa = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')     # Verde claro
        fill_programada = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Amarillo claro
        fill_cancelada = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')   # Naranja/Rojo claro
        
        # Título
        ws.merge_cells('A1:L1')
        cell_titulo = ws['A1']
        cell_titulo.value = "REPORTE DE PACKING / PRODUCCIÓN"
        cell_titulo.font = titulo_font
        cell_titulo.fill = titulo_fill
        cell_titulo.alignment = titulo_alignment
        ws.row_dimensions[1].height = 30
        
        # Encabezados
        headers = [
            'Empresa', 'Semana de Proyectado', 'Tipo Fruta', 
            'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado',
            'KG Total', 'Estado', 'Observaciones'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
        
        ws.row_dimensions[2].height = 35
        
        # Datos
        row_num = 3
        for packing in queryset:
            tipos = packing.tipos.all()
            
            # Si el packing no tiene tipos registrados, mostrar una fila base
            if not tipos:
                row_data = [
                    packing.empresa.nombre,
                    self._formato_semana(packing.fecha_inicio_semana, packing.fecha_fin_semana),
                    'Sin producción registrada',
                    0.0, 0.0, 0.0, 0.0, 0.0, 0.0,
                    0.0,
                    packing.get_estado_display(),
                    packing.observaciones or ''
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = cell_border
                    if 4 <= col_num <= 10:
                        cell.number_format = '#,##0.00'
                
                row_num += 1
                continue

            # Si tiene tipos, mostrar cada tipo
            for tipo in tipos:
                # Obtener kilos por día y los objetos PackingDetalle
                detalles_queryset = tipo.detalles.all()
                detalles_dict = {d.dia: d for d in detalles_queryset}
                
                # --- CALCULAR ESTADO ESPECÍFICO PARA ESTE TIPO (Optimizado) ---
                # Ahora usamos la property subasta_activa que retorna la subasta no cancelada
                subastas_tipo = []
                for detalle in detalles_queryset:
                    subasta = detalle.subasta_activa  # Property que retorna subasta no cancelada
                    if subasta:
                        subastas_tipo.append(subasta)
                
                estados_list = [s.estado_calculado for s in subastas_tipo]
                
                if not subastas_tipo:
                    estado_tipo = "PROYECTADO"
                else:
                    if 'ACTIVA' in estados_list:
                        estado_tipo = "EN SUBASTA"
                    elif any(e in ('FINALIZADA', 'CANCELADA') for e in estados_list):
                        # Se considera FINALIZADO si todos los días con producción tienen subastas 
                        # Y todas esas subastas están terminadas (Finalizada o Cancelada)
                        detalles_con_produccion = [d for d in detalles_queryset if d.py > 0]
                        detalles_con_subasta = [s.packing_detalle_id for s in subastas_tipo]
                        
                        todo_cubierto = all(d.id in detalles_con_subasta for d in detalles_con_produccion)
                        todas_terminadas = all(e in ('FINALIZADA', 'CANCELADA') for e in estados_list)
                        
                        if todo_cubierto and todas_terminadas:
                            estado_tipo = "FINALIZADO"
                        else:
                            estado_tipo = "PARCIAL"
                    elif 'PROGRAMADA' in estados_list:
                        estado_tipo = "PROGRAMADO"
                    else:
                        estado_tipo = "PROYECTADO"

                if packing.estado == 'ANULADO':
                    estado_tipo = "ANULADO"

                # Días de la semana para el mapeo
                dias_semana = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']
                
                row_data = [
                    packing.empresa.nombre,
                    self._formato_semana(packing.fecha_inicio_semana, packing.fecha_fin_semana),
                    tipo.tipo_fruta.nombre,
                ]
                # Llenar los kilos para cada día
                for d_nom in dias_semana:
                    row_data.append(float(detalles_dict.get(d_nom).py) if d_nom in detalles_dict else 0.0)
                
                # Agregar el resto de campos
                row_data.extend([
                    float(tipo.kg_total),
                    estado_tipo,
                    packing.observaciones or ''
                ])
                
                # Escribir la fila y aplicar color naranja SOLO a canceladas
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = cell_border
                    
                    # Formato numérico para KGs (Columnas D a J)
                    if 4 <= col_num <= 10:
                        cell.number_format = '#,##0.00'
                    
                    # --- APLICAR COLOR NARANJA SOLO SI EL DÍA ESTÁ CANCELADO SIN REACTIVAR ---
                    # Un día está "cancelado sin reactivar" si:
                    # - Tiene subastas (hay historial)
                    # - Pero NO tiene subasta activa (subasta_activa es None)
                    # Esto significa que solo tiene subastas CANCELADAS
                    if 4 <= col_num <= 9:
                        dia_nombre = dias_semana[col_num - 4]
                        detalle_dia = detalles_dict.get(dia_nombre)
                        
                        if detalle_dia and value > 0:
                            # subasta_activa retorna None si solo hay canceladas
                            subasta_activa = detalle_dia.subasta_activa
                            tiene_subastas = detalle_dia.subastas.exists()
                            
                            # Marcar naranja si tiene historial de subastas pero ninguna vigente
                            if tiene_subastas and subasta_activa is None:
                                cell.fill = fill_cancelada 
                
                row_num += 1
        
        # Anchos
        column_widths = {
            'A': 25, 'B': 35, 'C': 20, # B aumentado para "Semana X | dd/mm/yyyy..."
            'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 10,
            'J': 15, 'K': 15, 'L': 40
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        if row_num > 3:
            ws.auto_filter.ref = f"A2:L{row_num-1}"
            
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"reporte_packing_{timestamp}.xlsx"
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
